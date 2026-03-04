"""
Run this once to create:  gstr2b_import_sample.xlsx
Then upload that file to: POST /upload-excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GSTR-2B Import"

# ── Headers ──────────────────────────────────────────────────
headers = [
    "Party Name",
    "Invoice Number",
    "Invoice Date",
    "Taxable Value",
    "CGST",
    "SGST",
    "IGST",
]

header_fill   = PatternFill("solid", fgColor="1F4E79")
header_font   = Font(bold=True, color="FFFFFF", size=11)
header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = thin_border

# ── Sample data ───────────────────────────────────────────────
# Mix of B2B (CGST+SGST) and inter-state (IGST) invoices
# Party names must EXACTLY match ledger names in TallyPrime
rows = [
    # party,                  inv_no,       date,        taxable,  cgst,   sgst,   igst
    ("ABC TRADERS",           "INV-2025-01", "01-04-2025", 50000,  4500,   4500,   0),
    ("XYZ SUPPLIES PVT LTD",  "INV-2025-88", "15-04-2025", 120000, 10800,  10800,  0),
    ("SHARMA ENTERPRISES",    "SE/25-26/001","03-05-2025", 75000,  6750,   6750,   0),
    ("NATIONAL DISTRIBUTORS", "ND/2025/45",  "10-05-2025", 200000, 0,      0,      36000),  # IGST (inter-state)
    ("KUMAR & SONS",          "KS-100",      "22-05-2025", 30000,  2700,   2700,   0),
    ("MEHTA CHEMICALS LTD",   "MCL/25/789",  "01-06-2025", 95000,  0,      0,      17100),  # IGST
    ("RAJESH TRADERS",        "RT-2025-33",  "14-06-2025", 45000,  4050,   4050,   0),
    ("GLOBAL IMPORTS CO",     "GIC/2025/12", "30-06-2025", 180000, 0,      0,      32400),  # IGST
]

data_font    = Font(size=10)
data_align_l = Alignment(horizontal="left",   vertical="center")
data_align_r = Alignment(horizontal="right",  vertical="center")
data_align_c = Alignment(horizontal="center", vertical="center")

alt_fill = PatternFill("solid", fgColor="EBF3FB")  # light blue for alternate rows

for r_idx, row in enumerate(rows, start=2):
    fill = alt_fill if r_idx % 2 == 0 else PatternFill()
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font   = data_font
        cell.fill   = fill
        cell.border = thin_border
        if c_idx == 1:                          # Party Name
            cell.alignment = data_align_l
        elif c_idx in (2, 3):                   # Inv No, Date
            cell.alignment = data_align_c
        else:                                    # Numbers
            cell.alignment = data_align_r
            cell.number_format = '#,##0.00'

# ── Column widths ─────────────────────────────────────────────
col_widths = [30, 18, 14, 16, 12, 12, 12]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30

# ── Freeze header row ─────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Instructions sheet ────────────────────────────────────────
ws2 = wb.create_sheet("Instructions")
ws2.column_dimensions["A"].width = 90
instructions = [
    ("GSTR-2B Excel Import — Instructions", True),
    ("", False),
    ("REQUIRED columns (flexible naming — exact spelling not needed):", True),
    ("  • Party Name       — must exactly match the ledger name in TallyPrime", False),
    ("  • Invoice Number   — supplier's invoice number", False),
    ("  • Invoice Date     — any common date format: DD-MM-YYYY, DD/MM/YYYY, YYYY-MM-DD", False),
    ("  • Taxable Value    — taxable amount (excluding GST)", False),
    ("", False),
    ("OPTIONAL columns:", True),
    ("  • CGST   — Central GST amount (leave 0 for inter-state invoices)", False),
    ("  • SGST   — State GST amount   (leave 0 for inter-state invoices)", False),
    ("  • IGST   — Integrated GST     (leave 0 for intra-state invoices)", False),
    ("", False),
    ("STEPS:", True),
    ("  1. Fill in your invoice data in the 'GSTR-2B Import' sheet.", False),
    ("  2. POST /check-ledgers-excel  to find any missing ledgers in TallyPrime.", False),
    ("  3. Create missing ledgers as Sundry Creditors in TallyPrime.", False),
    ("  4. POST /upload-excel         to import all invoices.", False),
    ("", False),
    ("TIPS:", True),
    ("  • You can rename the columns — the tool matches by common aliases.", False),
    ("  • If CGST/SGST/IGST columns are omitted entirely, tax amounts default to 0.", False),
    ("  • Blank rows are skipped automatically.", False),
    ("  • CSV files work too — same column names apply.", False),
]
for r, (text, bold) in enumerate(instructions, start=1):
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=10 if not bold else 11)

wb.save("gstr2b_import_sample.xlsx")
print("✓  gstr2b_import_sample.xlsx created successfully.")
print("   Upload it to: POST http://localhost:8000/upload-excel")